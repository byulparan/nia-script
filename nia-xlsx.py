#!/usr/bin/env python

from openpyxl import load_workbook
from pathlib import Path
import re
import os

index = 1

# for file in Path('.').iterdir():
#     if file.suffix == '.xlsx':
#         new_type = re.compile('\d+-\d+-\d+').search(file.as_posix()).group()
#         wb = load_workbook(file)
#         ws = wb.active
#         value = ws.cell(row=3, column=3).value
#         if value :
#             print(index, file, '|', value, '|', new_type)
#             ws.cell(row=3, column=3, value= new_type)
#         else:
#             value = ws.cell(row=2, column=2).value
#             print(index, file, '|', value, '|', new_type)
#             ws.cell(row=2, column=2, value= new_type)
#         file.unlink()
#         wb.save(file.stem + '.xlsx')
#         index += 1

for file in Path('.').iterdir():
    if file.suffix == '.xlsx':
        wb = load_workbook(file)
        ws = wb.active
        value = ws.cell(row=10, column=3).value
        if value :
            print(index, file, '|', value, '|', "박성민")
            ws.cell(row=10, column=3, value= "박성민")
        else:
            value = ws.cell(row=4, column=2).value
            print(index, file, '|', value, '|', "박성민")
            ws.cell(row=4, column=2, value= "박성민")
        file.unlink()
        wb.save(file.stem + '.xlsx')
        index += 1


